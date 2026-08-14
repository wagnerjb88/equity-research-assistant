import yfinance as yf
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO

def get_stock_data(ticker):
    """
    Fetches a yfinance Ticker object and its info dictionary for a given ticker symbol.
    Returns (stock, info) or (None, None) if the ticker is invalid.
    """
    stock = yf.Ticker(ticker)
    info = stock.info

    # Some tickers populate longName, others only populate shortName.
    # Fall back through a few fields before concluding the ticker is invalid.
    name = info.get("longName") or info.get("shortName") or info.get("symbol")

    if not info or name is None:
        return None, None

    return stock, info
def get_price_history(stock, period="1y"):
    """
    Fetches historical price data for a given stock (yfinance Ticker object).
    period options: '1mo', '6mo', '1y', '5y', 'max', etc.
    """
    history = stock.history(period=period)
    return history
def get_financial_statements(stock, period="annual"):
    """
    Fetches income statement, balance sheet, and cash flow statement.
    period: 'annual' or 'quarterly'
    Returns a dict with three DataFrames.
    """
    if period == "annual":
        income_stmt = stock.financials
        balance_sheet = stock.balance_sheet
        cash_flow = stock.cashflow
    else:
        income_stmt = stock.quarterly_financials
        balance_sheet = stock.quarterly_balance_sheet
        cash_flow = stock.quarterly_cashflow

    return {
        "income_statement": income_stmt,
        "balance_sheet": balance_sheet,
        "cash_flow": cash_flow
    }
def get_key_metrics(info, stock):
    """
    Extracts and calculates key financial ratios/metrics from the info dict
    and cash flow statement.
    Returns a dict of grouped metrics.
    """
    # --- Free Cash Flow (calculated from cash flow statement) ---
    fcf = None
    fcf_margin = None
    try:
        cash_flow = stock.cashflow
        if not cash_flow.empty and "Free Cash Flow" in cash_flow.index:
            fcf = cash_flow.loc["Free Cash Flow"].iloc[0]
        elif not cash_flow.empty:
            # Fallback: Operating Cash Flow - CapEx
            op_cf = cash_flow.loc["Operating Cash Flow"].iloc[0] if "Operating Cash Flow" in cash_flow.index else None
            capex = cash_flow.loc["Capital Expenditure"].iloc[0] if "Capital Expenditure" in cash_flow.index else None
            if op_cf is not None and capex is not None:
                fcf = op_cf + capex  # capex is usually negative already

        total_revenue = info.get("totalRevenue")
        if fcf is not None and total_revenue:
            fcf_margin = fcf / total_revenue
    except Exception:
        fcf = None
        fcf_margin = None

    return {
        "valuation": {
            "P/E (TTM)": info.get("trailingPE"),
            "Forward P/E": info.get("forwardPE"),
            "P/B": info.get("priceToBook"),
            "P/S": info.get("priceToSalesTrailing12Months"),
            "EV/EBITDA": info.get("enterpriseToEbitda"),
        },
        "profitability": {
            "Gross Margin": info.get("grossMargins"),
            "Operating Margin": info.get("operatingMargins"),
            "Net Margin": info.get("profitMargins"),
            "ROE": info.get("returnOnEquity"),
            "ROA": info.get("returnOnAssets"),
        },
        "growth": {
            "Revenue Growth (YoY)": info.get("revenueGrowth"),
            "Earnings Growth (YoY)": info.get("earningsGrowth"),
        },
        "financial_health": {
            "Debt/Equity": info.get("debtToEquity"),
            "Current Ratio": info.get("currentRatio"),
            "Quick Ratio": info.get("quickRatio"),
        },
        "cash_generation": {
            "Free Cash Flow": fcf,
            "FCF Margin": fcf_margin,
        }
    }

def get_comparison_data(tickers):
    """
    Fetches key comparison metrics for a list of ticker symbols.
    Returns a list of dicts, one per company, skipping any that fail.
    """
    comparison_data = []

    for ticker in tickers:
        try:
            stock, info = get_stock_data(ticker)
            if stock is None:
                continue

            comparison_data.append({
                "Ticker": ticker,
                "Company": info.get("longName", ticker),
                "Market Cap": info.get("marketCap"),
                "P/E (TTM)": info.get("trailingPE"),
                "Forward P/E": info.get("forwardPE"),
                "P/B": info.get("priceToBook"),
                "EV/EBITDA": info.get("enterpriseToEbitda"),
                "Gross Margin": info.get("grossMargins"),
                "Operating Margin": info.get("operatingMargins"),
                "Net Margin": info.get("profitMargins"),
                "ROE": info.get("returnOnEquity"),
                "Revenue Growth": info.get("revenueGrowth"),
                "Debt/Equity": info.get("debtToEquity"),
            })
        except Exception:
            continue

    return comparison_data

SECTOR_THRESHOLDS = {
    "Technology": {
        "pe": {"good": 25, "bad": 60},
        "ev_ebitda": {"good": 15, "bad": 35},
        "net_margin": {"good": 0.20, "bad": 0.05},
        "revenue_growth": {"good": 0.15, "bad": 0.0},
    },
    "Financial Services": {
        "pe": {"good": 10, "bad": 20},
        "ev_ebitda": {"good": 8, "bad": 16},
        "net_margin": {"good": 0.25, "bad": 0.10},
        "revenue_growth": {"good": 0.08, "bad": 0.0},
    },
    "Healthcare": {
        "pe": {"good": 18, "bad": 40},
        "ev_ebitda": {"good": 12, "bad": 25},
        "net_margin": {"good": 0.15, "bad": 0.03},
        "revenue_growth": {"good": 0.10, "bad": 0.0},
    },
    "Consumer Cyclical": {
        "pe": {"good": 15, "bad": 30},
        "ev_ebitda": {"good": 10, "bad": 20},
        "net_margin": {"good": 0.08, "bad": 0.02},
        "revenue_growth": {"good": 0.08, "bad": 0.0},
    },
    "Consumer Defensive": {
        "pe": {"good": 15, "bad": 28},
        "ev_ebitda": {"good": 10, "bad": 18},
        "net_margin": {"good": 0.08, "bad": 0.02},
        "revenue_growth": {"good": 0.05, "bad": 0.0},
    },
    "Energy": {
        "pe": {"good": 10, "bad": 20},
        "ev_ebitda": {"good": 5, "bad": 12},
        "net_margin": {"good": 0.10, "bad": 0.02},
        "revenue_growth": {"good": 0.05, "bad": -0.05},
    },
    "Industrials": {
        "pe": {"good": 15, "bad": 28},
        "ev_ebitda": {"good": 10, "bad": 18},
        "net_margin": {"good": 0.10, "bad": 0.03},
        "revenue_growth": {"good": 0.06, "bad": 0.0},
    },
    "Utilities": {
        "pe": {"good": 16, "bad": 25},
        "ev_ebitda": {"good": 9, "bad": 14},
        "net_margin": {"good": 0.12, "bad": 0.05},
        "revenue_growth": {"good": 0.04, "bad": -0.02},
    },
    # Fallback for any sector not listed above
    "Default": {
        "pe": {"good": 15, "bad": 40},
        "ev_ebitda": {"good": 10, "bad": 25},
        "net_margin": {"good": 0.15, "bad": 0.02},
        "revenue_growth": {"good": 0.10, "bad": 0.0},
    },
}


def get_sector_thresholds(sector):
    """
    Returns the threshold dict for a given sector, falling back to Default
    if the sector isn't explicitly covered.
    """
    return SECTOR_THRESHOLDS.get(sector, SECTOR_THRESHOLDS["Default"])
def calculate_score(metrics, comparison_data=None, base_ticker=None, sector=None):
    """
    Calculates a 0-100 score for each category (Valuation, Profitability, Growth,
    Financial Health) plus an overall weighted score.
    Uses peer comparison if available, otherwise falls back to absolute thresholds.
    """
    use_peers = comparison_data is not None and len(comparison_data) > 1

    if use_peers:
        peer_df = pd.DataFrame(comparison_data).set_index("Ticker")

    def score_relative(metric_label, raw_value, lower_is_better=False):
        """Scores a metric 0-100 based on percentile rank among peers."""
        if raw_value is None or peer_df[metric_label].dropna().empty:
            return None
        values = peer_df[metric_label].dropna()
        if base_ticker not in values.index:
            return None
        rank = values.rank(pct=True, ascending=not lower_is_better)
        return round(rank.get(base_ticker, 0.5) * 100)

    def score_absolute(raw_value, good, bad, lower_is_better=False):
        """Scores a metric 0-100 based on fixed thresholds. 'good' and 'bad' define the scale."""
        if raw_value is None:
            return None
        if lower_is_better:
            if raw_value <= good:
                return 100
            if raw_value >= bad:
                return 0
            return round(100 * (bad - raw_value) / (bad - good))
        else:
            if raw_value >= good:
                return 100
            if raw_value <= bad:
                return 0
            return round(100 * (raw_value - bad) / (good - bad))

    def get_category_score(metric_scores):
        """Averages non-None scores in a category. Returns None if no data available."""
        valid_scores = [s for s in metric_scores if s is not None]
        if not valid_scores:
            return None
        return round(sum(valid_scores) / len(valid_scores))

    val = metrics["valuation"]
    prof = metrics["profitability"]
    growth = metrics["growth"]
    health = metrics["financial_health"]

    thresholds = get_sector_thresholds(sector)

    # --- Valuation (lower is generally "cheaper" = better) ---
    if use_peers:
        valuation_scores = [
            score_relative("P/E (TTM)", val["P/E (TTM)"], lower_is_better=True),
            score_relative("P/B", val["P/B"], lower_is_better=True),
            score_relative("EV/EBITDA", val["EV/EBITDA"], lower_is_better=True),
        ]
    else:
        valuation_scores = [
            score_absolute(val["P/E (TTM)"], good=thresholds["pe"]["good"], bad=thresholds["pe"]["bad"], lower_is_better=True),
            score_absolute(val["P/B"], good=2, bad=8, lower_is_better=True),
            score_absolute(val["EV/EBITDA"], good=thresholds["ev_ebitda"]["good"], bad=thresholds["ev_ebitda"]["bad"], lower_is_better=True),
        ]
    # --- Profitability (higher is better) ---
    if use_peers:
        profitability_scores = [
            score_relative("Gross Margin", prof["Gross Margin"]),
            score_relative("Operating Margin", prof["Operating Margin"]),
            score_relative("Net Margin", prof["Net Margin"]),
            score_relative("ROE", prof["ROE"]),
        ]
    else:
        profitability_scores = [
            score_absolute(prof["Gross Margin"], good=0.5, bad=0.15),
            score_absolute(prof["Operating Margin"], good=0.25, bad=0.05),
            score_absolute(prof["Net Margin"], good=thresholds["net_margin"]["good"], bad=thresholds["net_margin"]["bad"]),
            score_absolute(prof["ROE"], good=0.20, bad=0.05),
        ]

    # --- Growth (higher is better) ---
    if use_peers:
        growth_scores = [
            score_relative("Revenue Growth", growth["Revenue Growth (YoY)"]),
        ]
    else:
        growth_scores = [
            score_absolute(growth["Revenue Growth (YoY)"], good=thresholds["revenue_growth"]["good"], bad=thresholds["revenue_growth"]["bad"]),
            score_absolute(growth["Earnings Growth (YoY)"], good=0.15, bad=0.0),
        ]

    # --- Financial Health (lower debt / higher liquidity is better) ---
    if use_peers:
        health_scores = [
            score_relative("Debt/Equity", health["Debt/Equity"], lower_is_better=True),
        ]
    else:
        health_scores = [
            score_absolute(health["Debt/Equity"], good=50, bad=200, lower_is_better=True),
            score_absolute(health["Current Ratio"], good=2, bad=0.8),
            score_absolute(health["Quick Ratio"], good=1.5, bad=0.5),
        ]

    category_scores = {
        "Valuation": get_category_score(valuation_scores),
        "Profitability": get_category_score(profitability_scores),
        "Growth": get_category_score(growth_scores),
        "Financial Health": get_category_score(health_scores),
    }

    weights = {
        "Valuation": 0.20,
        "Profitability": 0.30,
        "Growth": 0.25,
        "Financial Health": 0.25,
    }

    weighted_total = 0
    weight_used = 0
    for category, score in category_scores.items():
        if score is not None:
            weighted_total += score * weights[category]
            weight_used += weights[category]

    overall_score = round(weighted_total / weight_used) if weight_used > 0 else None

    return {
        "categories": category_scores,
        "overall": overall_score,
        "method": "peer-relative" if use_peers else "absolute thresholds"
    }

def generate_score_explanation(metrics, score_result):
    """
    Generates plain-English bullet points explaining why each category scored as it did.
    Returns a dict: category -> list of explanation strings.
    """
    explanations = {}

    val = metrics["valuation"]
    prof = metrics["profitability"]
    growth = metrics["growth"]
    health = metrics["financial_health"]

    # --- Valuation ---
    notes = []
    pe = val["P/E (TTM)"]
    if pe is not None:
        if pe < 15:
            notes.append(f"P/E of {pe:.1f} is low, suggesting the stock may be undervalued.")
        elif pe > 35:
            notes.append(f"P/E of {pe:.1f} is elevated, suggesting the market is pricing in high expectations.")
        else:
            notes.append(f"P/E of {pe:.1f} is in a moderate range.")
    ev_ebitda = val["EV/EBITDA"]
    if ev_ebitda is not None:
        if ev_ebitda > 20:
            notes.append(f"EV/EBITDA of {ev_ebitda:.1f} is on the higher side relative to typical benchmarks.")
        elif ev_ebitda < 10:
            notes.append(f"EV/EBITDA of {ev_ebitda:.1f} looks attractive relative to typical benchmarks.")
    explanations["Valuation"] = notes

    # --- Profitability ---
    notes = []
    net_margin = prof["Net Margin"]
    if net_margin is not None:
        if net_margin > 0.20:
            notes.append(f"Net margin of {net_margin*100:.1f}% is strong, indicating efficient conversion of revenue to profit.")
        elif net_margin < 0.05:
            notes.append(f"Net margin of {net_margin*100:.1f}% is thin, indicating limited profit conversion.")
    roe = prof["ROE"]
    if roe is not None:
        if roe > 0.20:
            notes.append(f"ROE of {roe*100:.1f}% shows strong returns generated on shareholder equity.")
        elif roe < 0.08:
            notes.append(f"ROE of {roe*100:.1f}% is relatively weak.")
    explanations["Profitability"] = notes

    # --- Growth ---
    notes = []
    rev_growth = growth["Revenue Growth (YoY)"]
    if rev_growth is not None:
        if rev_growth > 0.15:
            notes.append(f"Revenue growth of {rev_growth*100:.1f}% YoY reflects strong top-line momentum.")
        elif rev_growth < 0:
            notes.append(f"Revenue declined {abs(rev_growth)*100:.1f}% YoY, a red flag worth investigating.")
        else:
            notes.append(f"Revenue growth of {rev_growth*100:.1f}% YoY is modest.")
    explanations["Growth"] = notes

    # --- Financial Health ---
    notes = []
    debt_equity = health["Debt/Equity"]
    if debt_equity is not None:
        if debt_equity > 150:
            notes.append(f"Debt/Equity of {debt_equity:.0f} is elevated, indicating meaningful leverage risk.")
        elif debt_equity < 50:
            notes.append(f"Debt/Equity of {debt_equity:.0f} is conservative, indicating low leverage risk.")
    current_ratio = health["Current Ratio"]
    if current_ratio is not None:
        if current_ratio < 1:
            notes.append(f"Current ratio of {current_ratio:.2f} is below 1, meaning short-term liabilities exceed short-term assets.")
        elif current_ratio > 2:
            notes.append(f"Current ratio of {current_ratio:.2f} indicates strong short-term liquidity.")
    explanations["Financial Health"] = notes

    return explanations
def calculate_comps_valuation(info, comparison_data):
    """
    Estimates fair value using peer median valuation multiples (P/E, EV/EBITDA, P/S, P/B)
    applied to the company's own financials.
    Returns implied prices per method, plus per-peer multiple detail for transparency.
    """
    if not comparison_data or len(comparison_data) < 2:
        return None

    peer_df = pd.DataFrame(comparison_data).set_index("Ticker")
    base_ticker = info.get("symbol")

    if base_ticker not in peer_df.index:
        return None

    peers_only = peer_df.drop(index=base_ticker, errors="ignore")

    current_price = info.get("currentPrice")
    shares_outstanding = info.get("sharesOutstanding")
    eps = info.get("trailingEps")
    ebitda = info.get("ebitda")
    total_debt = info.get("totalDebt") or 0
    cash = info.get("totalCash") or 0
    revenue_per_share = info.get("totalRevenue") / shares_outstanding if info.get("totalRevenue") and shares_outstanding else None
    book_value_per_share = info.get("bookValue")

    results = {}

    def clean_multiples(series):
        """Drops missing and negative/extreme multiples that would distort the median."""
        return series.dropna()[(series > 0) & (series < 200)]

    # --- P/E Method ---
    pe_values = clean_multiples(peers_only["P/E (TTM)"])
    if not pe_values.empty and eps is not None and eps > 0:
        median_pe = pe_values.median()
        implied_price = median_pe * eps
        results["P/E"] = {
            "peer_median_multiple": round(median_pe, 2),
            "peer_multiples": pe_values.round(2).to_dict(),
            "implied_price": round(implied_price, 2),
        }

    # --- EV/EBITDA Method ---
    ev_ebitda_values = clean_multiples(peers_only["EV/EBITDA"])
    if not ev_ebitda_values.empty and ebitda is not None and ebitda > 0 and shares_outstanding:
        median_ev_ebitda = ev_ebitda_values.median()
        implied_ev = median_ev_ebitda * ebitda
        implied_equity_value = implied_ev - total_debt + cash
        implied_price = implied_equity_value / shares_outstanding
        results["EV/EBITDA"] = {
            "peer_median_multiple": round(median_ev_ebitda, 2),
            "peer_multiples": ev_ebitda_values.round(2).to_dict(),
            "implied_price": round(implied_price, 2),
        }

    # --- P/S Method ---
    if "P/S" in peers_only.columns:
        ps_values = clean_multiples(peers_only["P/S"])
    else:
        ps_values = pd.Series(dtype=float)
    if not ps_values.empty and revenue_per_share is not None and revenue_per_share > 0:
        median_ps = ps_values.median()
        implied_price = median_ps * revenue_per_share
        results["P/S"] = {
            "peer_median_multiple": round(median_ps, 2),
            "peer_multiples": ps_values.round(2).to_dict(),
            "implied_price": round(implied_price, 2),
        }

    # --- P/B Method ---
    if "P/B" in peers_only.columns:
        pb_values = clean_multiples(peers_only["P/B"])
    else:
        pb_values = pd.Series(dtype=float)
    if not pb_values.empty and book_value_per_share is not None and book_value_per_share > 0:
        median_pb = pb_values.median()
        implied_price = median_pb * book_value_per_share
        results["P/B"] = {
            "peer_median_multiple": round(median_pb, 2),
            "peer_multiples": pb_values.round(2).to_dict(),
            "implied_price": round(implied_price, 2),
        }

    if not results:
        return None

    implied_prices = [r["implied_price"] for r in results.values()]
    low = min(implied_prices)
    high = max(implied_prices)
    avg_price = sum(implied_prices) / len(implied_prices)

    return {
        "methods": results,
        "low": round(low, 2),
        "high": round(high, 2),
        "average_implied_price": round(avg_price, 2),
        "current_price": current_price,
        "average_upside_pct": round((avg_price - current_price) / current_price * 100, 1) if current_price else None,
    }

    # Exclude the base company from the peer average
    peers_only = peer_df.drop(index=base_ticker, errors="ignore")

    current_price = info.get("currentPrice")
    shares_outstanding = info.get("sharesOutstanding")
    eps = info.get("trailingEps")
    ebitda = info.get("ebitda")
    total_debt = info.get("totalDebt") or 0
    cash = info.get("totalCash") or 0

    results = {}

    # --- P/E-based valuation ---
    peer_pe_avg = peers_only["P/E (TTM)"].dropna().mean()
    if pd.notna(peer_pe_avg) and eps is not None and eps > 0:
        implied_price_pe = peer_pe_avg * eps
        results["P/E Method"] = {
            "peer_avg_multiple": round(peer_pe_avg, 2),
            "implied_price": round(implied_price_pe, 2),
            "upside_pct": round((implied_price_pe - current_price) / current_price * 100, 1) if current_price else None
        }

    # --- EV/EBITDA-based valuation ---
    peer_ev_ebitda_avg = peers_only["EV/EBITDA"].dropna().mean()
    if pd.notna(peer_ev_ebitda_avg) and ebitda is not None and ebitda > 0 and shares_outstanding:
        implied_ev = peer_ev_ebitda_avg * ebitda
        implied_equity_value = implied_ev - total_debt + cash
        implied_price_ev = implied_equity_value / shares_outstanding
        results["EV/EBITDA Method"] = {
            "peer_avg_multiple": round(peer_ev_ebitda_avg, 2),
            "implied_price": round(implied_price_ev, 2),
            "upside_pct": round((implied_price_ev - current_price) / current_price * 100, 1) if current_price else None
        }

    if not results:
        return None

    # Average across whichever methods produced a result
    avg_implied_price = sum(r["implied_price"] for r in results.values()) / len(results)

    return {
        "methods": results,
        "average_implied_price": round(avg_implied_price, 2),
        "current_price": current_price,
        "average_upside_pct": round((avg_implied_price - current_price) / current_price * 100, 1) if current_price else None
    }
def calculate_dcf(info, stock, growth_rate=None, discount_rate=None, terminal_growth=None, projection_years=5):
    """
    Calculates intrinsic value per share using a Discounted Cash Flow model.
    If growth_rate/discount_rate/terminal_growth are not provided, reasonable
    defaults are derived from the company's own data.
    """
    shares_outstanding = info.get("sharesOutstanding")
    total_debt = info.get("totalDebt") or 0
    cash = info.get("totalCash") or 0
    beta = info.get("beta") or 1.0
    current_price = info.get("currentPrice")

    # --- Get starting Free Cash Flow ---
    try:
        cash_flow = stock.cashflow
        if not cash_flow.empty and "Free Cash Flow" in cash_flow.index:
            starting_fcf = cash_flow.loc["Free Cash Flow"].iloc[0]
        else:
            op_cf = cash_flow.loc["Operating Cash Flow"].iloc[0] if "Operating Cash Flow" in cash_flow.index else None
            capex = cash_flow.loc["Capital Expenditure"].iloc[0] if "Capital Expenditure" in cash_flow.index else None
            starting_fcf = (op_cf + capex) if (op_cf is not None and capex is not None) else None
    except Exception:
        starting_fcf = None

    if starting_fcf is None or starting_fcf <= 0 or shares_outstanding is None:
        return None

    # --- Auto-calculate default assumptions if not provided ---
    if growth_rate is None:
        rev_growth = info.get("revenueGrowth")
        # Use historical revenue growth as a proxy, capped to a sane range
        growth_rate = rev_growth if rev_growth is not None else 0.08
        growth_rate = max(min(growth_rate, 0.25), -0.05)  # cap between -5% and 25%

    if discount_rate is None:
        # Simplified CAPM: risk-free rate + beta * equity risk premium
        risk_free_rate = 0.04
        equity_risk_premium = 0.05
        discount_rate = risk_free_rate + beta * equity_risk_premium
        discount_rate = max(min(discount_rate, 0.15), 0.06)  # cap between 6% and 15%

    if terminal_growth is None:
        terminal_growth = 0.025  # long-run GDP-like growth

    # --- Project future FCF ---
    projected_fcf = []
    fcf = starting_fcf
    for year in range(1, projection_years + 1):
        fcf = fcf * (1 + growth_rate)
        projected_fcf.append(fcf)

    # --- Discount each year's FCF back to present value ---
    discounted_fcf = [
        fcf / ((1 + discount_rate) ** year)
        for year, fcf in enumerate(projected_fcf, start=1)
    ]

    # --- Terminal value (Gordon Growth Model), discounted back to present ---
    terminal_value = (projected_fcf[-1] * (1 + terminal_growth)) / (discount_rate - terminal_growth)
    discounted_terminal_value = terminal_value / ((1 + discount_rate) ** projection_years)

    # --- Enterprise Value = sum of discounted FCF + discounted terminal value ---
    enterprise_value = sum(discounted_fcf) + discounted_terminal_value

    # --- Bridge to equity value and per-share price ---
    equity_value = enterprise_value - total_debt + cash
    implied_price = equity_value / shares_outstanding

    upside_pct = round((implied_price - current_price) / current_price * 100, 1) if current_price else None

    return {
        "assumptions": {
            "growth_rate": round(growth_rate * 100, 2),
            "discount_rate": round(discount_rate * 100, 2),
            "terminal_growth": round(terminal_growth * 100, 2),
            "projection_years": projection_years,
        },
        "starting_fcf": round(starting_fcf, 0),
        "projected_fcf": [round(f, 0) for f in projected_fcf],
        "discounted_fcf": [round(f, 0) for f in discounted_fcf],
        "terminal_value": round(terminal_value, 0),
        "discounted_terminal_value": round(discounted_terminal_value, 0),
        "enterprise_value": round(enterprise_value, 0),
        "equity_value": round(equity_value, 0),
        "implied_price": round(implied_price, 2),
        "current_price": current_price,
        "upside_pct": upside_pct,
    }
def generate_dcf_excel(info, dcf_result, ticker):
    """
    Builds a downloadable Excel DCF model with live formulas (not hardcoded values),
    so the user can adjust assumptions directly in Excel and see it recalculate.
    Returns an in-memory BytesIO buffer containing the .xlsx file.
    """
    company_name = info.get("longName") or info.get("shortName") or ticker
    shares_outstanding = info.get("sharesOutstanding")
    total_debt = info.get("totalDebt") or 0
    cash = info.get("totalCash") or 0
    current_price = info.get("currentPrice")

    a = dcf_result["assumptions"]
    starting_fcf = dcf_result["starting_fcf"]
    growth_rate = a["growth_rate"] / 100
    discount_rate = a["discount_rate"] / 100
    terminal_growth = a["terminal_growth"] / 100
    projection_years = a["projection_years"]

    wb = Workbook()
    ws = wb.active
    ws.title = "DCF Model"

    BOLD = Font(bold=True)
    BOLD_BLUE = Font(bold=True, color="0000FF")
    HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    GREEN_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    CURRENCY_FMT = '$#,##0;($#,##0);"-"'
    PERCENT_FMT = '0.0%'
    PRICE_FMT = '$#,##0.00'

    # --- Title ---
    ws["B2"] = f"{company_name} ({ticker}) — DCF Valuation Model"
    ws["B2"].font = Font(bold=True, size=14)
    ws.merge_cells("B2:G2")

    ws["B3"] = "All dollar figures in USD unless noted. Blue cells are editable inputs."
    ws["B3"].font = Font(italic=True, size=9, color="666666")
    ws.merge_cells("B3:G3")

    # --- Assumptions block ---
    ws["B5"] = "Key Assumptions"
    ws["B5"].fill = HEADER_FILL
    ws["B5"].font = HEADER_FONT
    ws.merge_cells("B5:C5")

    assumptions_list = [
        ("Starting Free Cash Flow ($)", starting_fcf, CURRENCY_FMT),
        ("FCF Growth Rate (Yr 1-5)", growth_rate, PERCENT_FMT),
        ("Discount Rate (WACC/CAPM)", discount_rate, PERCENT_FMT),
        ("Terminal Growth Rate", terminal_growth, PERCENT_FMT),
        ("Shares Outstanding", shares_outstanding, '#,##0'),
        ("Total Debt ($)", total_debt, CURRENCY_FMT),
        ("Cash & Equivalents ($)", cash, CURRENCY_FMT),
        ("Current Share Price ($)", current_price, PRICE_FMT),
    ]

    row = 6
    for label, value, fmt in assumptions_list:
        ws[f"B{row}"] = label
        ws[f"C{row}"] = value
        ws[f"C{row}"].font = BOLD_BLUE
        ws[f"C{row}"].fill = YELLOW_FILL
        ws[f"C{row}"].number_format = fmt
        row += 1

    CELL_START_FCF = "$C$6"
    CELL_GROWTH = "$C$7"
    CELL_DISCOUNT = "$C$8"
    CELL_TERMINAL_GROWTH = "$C$9"
    CELL_SHARES = "$C$10"
    CELL_DEBT = "$C$11"
    CELL_CASH = "$C$12"
    CELL_PRICE = "$C$13"

    # --- Projection table ---
    proj_header_row = row + 2
    ws[f"B{proj_header_row}"] = "Free Cash Flow Projection"
    ws[f"B{proj_header_row}"].fill = HEADER_FILL
    ws[f"B{proj_header_row}"].font = HEADER_FONT
    ws.merge_cells(f"B{proj_header_row}:G{proj_header_row}")

    col_header_row = proj_header_row + 1
    headers = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for i, h in enumerate(headers):
        col = get_column_letter(2 + i)
        ws[f"{col}{col_header_row}"] = h
        ws[f"{col}{col_header_row}"].font = BOLD

    fcf_row = col_header_row + 1
    ws[f"B{fcf_row}"] = "Projected FCF"
    for i in range(projection_years):
        col = get_column_letter(3 + i)
        if i == 0:
            formula = f"={CELL_START_FCF}*(1+{CELL_GROWTH})"
        else:
            prev_col = get_column_letter(3 + i - 1)
            formula = f"={prev_col}{fcf_row}*(1+{CELL_GROWTH})"
        ws[f"{col}{fcf_row}"] = formula
        ws[f"{col}{fcf_row}"].number_format = CURRENCY_FMT

    discount_factor_row = fcf_row + 1
    ws[f"B{discount_factor_row}"] = "Discount Factor"
    for i in range(projection_years):
        col = get_column_letter(3 + i)
        year_num = i + 1
        ws[f"{col}{discount_factor_row}"] = f"=1/(1+{CELL_DISCOUNT})^{year_num}"
        ws[f"{col}{discount_factor_row}"].number_format = '0.000'

    pv_fcf_row = discount_factor_row + 1
    ws[f"B{pv_fcf_row}"] = "PV of FCF"
    for i in range(projection_years):
        col = get_column_letter(3 + i)
        ws[f"{col}{pv_fcf_row}"] = f"={col}{fcf_row}*{col}{discount_factor_row}"
        ws[f"{col}{pv_fcf_row}"].number_format = CURRENCY_FMT
        ws[f"{col}{pv_fcf_row}"].font = BOLD

    # --- Terminal Value ---
    tv_row = pv_fcf_row + 2
    ws[f"B{tv_row}"] = "Terminal Value (Gordon Growth)"
    last_fcf_col = get_column_letter(2 + projection_years)
    ws[f"C{tv_row}"] = f"={last_fcf_col}{fcf_row}*(1+{CELL_TERMINAL_GROWTH})/({CELL_DISCOUNT}-{CELL_TERMINAL_GROWTH})"
    ws[f"C{tv_row}"].number_format = CURRENCY_FMT

    pv_tv_row = tv_row + 1
    ws[f"B{pv_tv_row}"] = "PV of Terminal Value"
    last_discount_col = get_column_letter(2 + projection_years)
    ws[f"C{pv_tv_row}"] = f"=C{tv_row}*{last_discount_col}{discount_factor_row}"
    ws[f"C{pv_tv_row}"].number_format = CURRENCY_FMT
    ws[f"C{pv_tv_row}"].font = BOLD

    # --- Valuation Bridge ---
    bridge_row = pv_tv_row + 2
    ws[f"B{bridge_row}"] = "Valuation Bridge"
    ws[f"B{bridge_row}"].fill = HEADER_FILL
    ws[f"B{bridge_row}"].font = HEADER_FONT
    ws.merge_cells(f"B{bridge_row}:C{bridge_row}")

    sum_pv_row = bridge_row + 1
    ws[f"B{sum_pv_row}"] = "Sum of PV of FCF (Yr 1-5)"
    last_pv_col = get_column_letter(2 + projection_years)
    ws[f"C{sum_pv_row}"] = f"=SUM(C{pv_fcf_row}:{last_pv_col}{pv_fcf_row})"
    ws[f"C{sum_pv_row}"].number_format = CURRENCY_FMT

    pv_tv_ref_row = sum_pv_row + 1
    ws[f"B{pv_tv_ref_row}"] = "PV of Terminal Value"
    ws[f"C{pv_tv_ref_row}"] = f"=C{pv_tv_row}"
    ws[f"C{pv_tv_ref_row}"].number_format = CURRENCY_FMT

    ev_row = pv_tv_ref_row + 1
    ws[f"B{ev_row}"] = "Enterprise Value"
    ws[f"C{ev_row}"] = f"=C{sum_pv_row}+C{pv_tv_ref_row}"
    ws[f"C{ev_row}"].number_format = CURRENCY_FMT
    ws[f"C{ev_row}"].font = BOLD

    less_debt_row = ev_row + 1
    ws[f"B{less_debt_row}"] = "Less: Total Debt"
    ws[f"C{less_debt_row}"] = f"=-{CELL_DEBT}"
    ws[f"C{less_debt_row}"].number_format = CURRENCY_FMT

    plus_cash_row = less_debt_row + 1
    ws[f"B{plus_cash_row}"] = "Plus: Cash & Equivalents"
    ws[f"C{plus_cash_row}"] = f"={CELL_CASH}"
    ws[f"C{plus_cash_row}"].number_format = CURRENCY_FMT

    equity_value_row = plus_cash_row + 1
    ws[f"B{equity_value_row}"] = "Equity Value"
    ws[f"C{equity_value_row}"] = f"=C{ev_row}+C{less_debt_row}+C{plus_cash_row}"
    ws[f"C{equity_value_row}"].number_format = CURRENCY_FMT
    ws[f"C{equity_value_row}"].font = BOLD

    shares_row = equity_value_row + 1
    ws[f"B{shares_row}"] = "Shares Outstanding"
    ws[f"C{shares_row}"] = f"={CELL_SHARES}"
    ws[f"C{shares_row}"].number_format = '#,##0'

    implied_price_row = shares_row + 1
    ws[f"B{implied_price_row}"] = "Implied Price per Share"
    ws[f"C{implied_price_row}"] = f"=C{equity_value_row}/C{shares_row}"
    ws[f"C{implied_price_row}"].number_format = PRICE_FMT
    ws[f"C{implied_price_row}"].font = Font(bold=True, size=12)
    ws[f"C{implied_price_row}"].fill = GREEN_FILL

    current_price_row = implied_price_row + 1
    ws[f"B{current_price_row}"] = "Current Share Price"
    ws[f"C{current_price_row}"] = f"={CELL_PRICE}"
    ws[f"C{current_price_row}"].number_format = PRICE_FMT

    upside_row = current_price_row + 1
    ws[f"B{upside_row}"] = "Implied Upside / (Downside)"
    ws[f"C{upside_row}"] = f"=(C{implied_price_row}-C{current_price_row})/C{current_price_row}"
    ws[f"C{upside_row}"].number_format = PERCENT_FMT
    ws[f"C{upside_row}"].font = BOLD

    # --- Column widths ---
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    # --- Save to in-memory buffer instead of disk ---
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_investment_thesis(info, score_result, dcf_result, comps_result, metrics, ticker):
    """
    Generates a plain-English investment thesis paragraph and a recommendation label,
    based on the company's score and valuation signals.
    """
    company_name = info.get("longName") or info.get("shortName") or ticker
    sector = info.get("sector", "its sector")
    industry = info.get("industry", "")
    overall_score = score_result["overall"] if score_result else None

    # --- Gather upside signals from whichever valuation methods are available ---
    upside_signals = []
    if dcf_result is not None and dcf_result.get("upside_pct") is not None:
        upside_signals.append(dcf_result["upside_pct"])
    if comps_result is not None and comps_result.get("average_upside_pct") is not None:
        upside_signals.append(comps_result["average_upside_pct"])

    avg_upside = sum(upside_signals) / len(upside_signals) if upside_signals else None

    # --- Determine recommendation label ---
    if overall_score is not None and avg_upside is not None:
        if overall_score >= 65 and avg_upside >= 10:
            recommendation = "Attractive"
        elif overall_score <= 40 or avg_upside <= -15:
            recommendation = "Unattractive"
        else:
            recommendation = "Neutral / Hold"
    elif overall_score is not None:
        recommendation = "Attractive" if overall_score >= 65 else "Unattractive" if overall_score <= 40 else "Neutral / Hold"
    else:
        recommendation = "Insufficient Data"

    sentences = []

    # --- Opening: company + sector context ---
    industry_str = f" within the {industry} industry" if industry else ""
    sentences.append(f"{company_name} ({ticker}) operates in the {sector} sector{industry_str}.")

    # --- Overall score ---
    if overall_score is not None:
        quality_desc = "strong" if overall_score >= 65 else "weak" if overall_score <= 40 else "mixed"
        sentences.append(f"The company scores {overall_score}/100 overall on our composite framework, reflecting {quality_desc} fundamentals across valuation, profitability, growth, and financial health.")
    else:
        sentences.append("Insufficient data was available to generate a composite fundamental score.")

    # --- Category strengths/weaknesses with specifics ---
    if score_result is not None:
        categories = score_result["categories"]
        valid_categories = {k: v for k, v in categories.items() if v is not None}
        if valid_categories:
            strongest = max(valid_categories, key=valid_categories.get)
            weakest = min(valid_categories, key=valid_categories.get)
            if strongest != weakest:
                sentences.append(f"{strongest} is the standout category at {categories[strongest]}/100, while {weakest} lags at {categories[weakest]}/100 and represents the primary area of concern.")

    # --- Profitability/growth specifics, pulled from metrics if available ---
    if metrics is not None:
        prof = metrics.get("profitability", {})
        growth = metrics.get("growth", {})
        health = metrics.get("financial_health", {})

        detail_notes = []
        net_margin = prof.get("Net Margin")
        if net_margin is not None:
            detail_notes.append(f"a net margin of {net_margin*100:.1f}%")
        rev_growth = growth.get("Revenue Growth (YoY)")
        if rev_growth is not None:
            detail_notes.append(f"YoY revenue growth of {rev_growth*100:.1f}%")
        debt_equity = health.get("Debt/Equity")
        if debt_equity is not None:
            leverage_desc = "elevated" if debt_equity > 150 else "conservative" if debt_equity < 50 else "moderate"
            detail_notes.append(f"{leverage_desc} leverage (D/E of {debt_equity:.0f})")

        if detail_notes:
            detail_str = ", ".join(detail_notes[:-1]) + (f", and {detail_notes[-1]}" if len(detail_notes) > 1 else detail_notes[0])
            sentences.append(f"On the fundamentals, the company reports {detail_str}.")

    # --- Valuation methodology and conclusion ---
    methods_used = []
    if dcf_result is not None and dcf_result.get("upside_pct") is not None:
        methods_used.append(f"a DCF model implies {dcf_result['upside_pct']:+.1f}% {'upside' if dcf_result['upside_pct'] > 0 else 'downside'} (${dcf_result['implied_price']:.2f} implied vs. ${dcf_result['current_price']:.2f} current)")
    if comps_result is not None and comps_result.get("average_upside_pct") is not None:
        methods_used.append(f"comparable company analysis implies {comps_result['average_upside_pct']:+.1f}% {'upside' if comps_result['average_upside_pct'] > 0 else 'downside'} (${comps_result['average_implied_price']:.2f} implied)")

    if methods_used:
        methods_str = " and ".join(methods_used)
        sentences.append(f"On valuation, {methods_str}.")
    else:
        sentences.append("No valuation estimate could be calculated — peer tickers and/or sufficient cash flow data are required for DCF and comps analysis.")

    # --- Closing recommendation sentence ---
    closing = {
        "Attractive": f"Taken together, {ticker} screens as an attractive investment candidate warranting further diligence.",
        "Neutral / Hold": f"Taken together, {ticker} presents a balanced risk/reward profile without a clear directional edge at current levels.",
        "Unattractive": f"Taken together, {ticker} screens as unattractive at current levels given weak fundamentals and/or limited valuation upside.",
        "Insufficient Data": f"Additional data is needed before forming a complete view on {ticker}.",
    }
    sentences.append(closing.get(recommendation, ""))

    thesis_paragraph = " ".join(s for s in sentences if s)

    return {
        "recommendation": recommendation,
        "thesis_paragraph": thesis_paragraph,
        "overall_score": overall_score,
        "avg_upside_pct": round(avg_upside, 1) if avg_upside is not None else None,
    }